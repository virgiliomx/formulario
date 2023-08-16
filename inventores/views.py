from django.http import HttpResponseRedirect
from django.shortcuts import redirect, render, get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse

import openpyxl
from .models import Inventor, Correo, Telefono
from .forms import DependenciasForm, InventorForm, CorreoForm, TelefonoForm, ProyectoForm


def nuevaDependencia(request):
    if request.method == 'POST':
        form = DependenciasForm(request.POST)
        if form.is_valid():
            form.save()
    else:
        form = DependenciasForm()
    return render(request, 'inventores/nuevaDependencia.html', {'form': form})


def cargaExcel(request):

    def dat(value):
        if value == "":
            return None
        return value

    if 'GET' == request.method:
        return render(request, 'inventores/cargaExcel.html', {})
    else:
        excel_file = request.FILES['excel_file']

        wb = openpyxl.load_workbook(excel_file)
        hoja = wb.active
        excel_data = []
        for i in range(2, hoja.max_row+1):
            nombre = dat(hoja.cell(i, 1).value)
            tipo = dat(hoja.cell(i, 2).value)
            numero = dat(hoja.cell(i, 3).value)
            dependencia = dat(hoja.cell(i, 4).value)

            inventor = Inventor(nombre=nombre, tipo=tipo,
                                numero=numero, dependencia=dependencia)
            inventor.save()
        return redirect(reverse('inventores:listaInventores'))
    return render(request, 'inventores/cargaExcel.html')


def listaInventores(request):
    inventores = Inventor.objects.all()
    paginator = Paginator(inventores, 20)
    page_number = request.GET.get('page', 1)
    try:
        inventores = paginator.page(page_number)
    except PageNotAnInteger:
        inventores = paginator.page(1)
    except EmptyPage:
        inventores = paginator.page(paginator.num_pages)
    return render(request, 'inventores/index.html', {'inventores': inventores})


def detalleInventor(request, slug):
    inventor = get_object_or_404(Inventor, slug=slug)
    correos = Correo.objects.filter(inventor=inventor)
    telefonos = Telefono.objects.filter(inventor=inventor)
    registros = inventor.registros.all()
    return render(request, 'inventores/detalleInventor.html', {'inventor': inventor, 'correos': correos, 'telefonos': telefonos, 'registros': registros})


def nuevoInventor(request):
    if request.method == 'POST':
        form = InventorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inventores:listaInventores')
    else:
        form = InventorForm()
    return render(request, 'inventores/nuevo.html', {'form': form})


def nuevoCorreo(request, slug):
    inventor = get_object_or_404(Inventor, slug=slug)
    if request.method == 'POST':
        form = CorreoForm(request.POST)
        if form.is_valid():
            correo = form.save(commit=False)
            correo.inventor = inventor
            form.save()
            return redirect('inventores:detalleInventor', inventor.slug)
    else:
        form = CorreoForm()
    return render(request, 'inventores/nuevoCorreo.html', {'form': form, 'inventor': inventor})


def nuevoTelefono(request, slug):
    inventor = get_object_or_404(Inventor, slug=slug)
    if request.method == 'POST':
        form = TelefonoForm(request.POST)
        if form.is_valid():
            telefono = form.save(commit=False)
            telefono.inventor = inventor
            form.save()
            return redirect('inventores:detalleInventor', inventor.slug)
    else:
        form = TelefonoForm()
    return render(request, 'inventores/nuevoTelefono.html', {'form': form, 'inventor': inventor})


def borraInventor(request, slug):
    inventor = get_object_or_404(Inventor, slug=slug)
    inventor.delete()
    return HttpResponseRedirect(reverse('inventores:listaInventores'))


def borraCorreo(request, correo_id, slug):
    correo = get_object_or_404(Correo, pk=correo_id)
    inventor = get_object_or_404(Inventor, slug=slug)
    correo.delete()
    return render(request, 'inventores/detalleInventor.html', {'inventor': inventor})


def borraTelefono(request, telefono_id, slug):
    telefono = get_object_or_404(Telefono, pk=telefono_id)
    inventor = get_object_or_404(Inventor, slug=slug)
    telefono.delete()
    return render(request, 'inventores/detalleInventor.html', {'inventor': inventor})


def nuevoRegistro(request, slug):
    inventor = get_object_or_404(Inventor, slug=slug)
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=inventor)
        if form.is_valid():
            form.save()
            return redirect('inventores:detalleInventor', inventor.slug)
    else:
        form = ProyectoForm()
    return render(request, 'inventores/nuevoProyecto.html', {'form': form, 'inventor': inventor})
